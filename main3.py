import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from Functions3 import get_files, map_files_with_keys, extracting_files_names_from_excel, border_column, border_top_cell, find_incorrect_file_name, check_incorrect_item

files = get_files()

files_dict = map_files_with_keys(files)

file_number = int(input('Please select file: >'))

list_files_names = extracting_files_names_from_excel(files_dict, file_number)


# name_format = {1:4, 2:3, 3:2, 4:2, 5:2, 6:3, 7:4, 8:range(1, 100)}
incorrect_names = []
incorrect_items = []
incorrect_items_code = []
for file_name in list_files_names:
    find_incorrect_file_name(file_name, incorrect_names, incorrect_items, incorrect_items_code)

# print(incorrect_names)
# print(incorrect_items)

wb = Workbook()
ws = wb.active
ws['A1'] = 'File Names'
ws['B1'] = 'Incorrect Item'
ws['C1'] = 'Project Code'
ws['D1'] = 'Originator'
ws['E1'] = 'Zone/ System'
ws['F1'] = 'Level'
ws['G1'] = 'Type'
ws['H1'] = 'Role'
ws['I1'] = 'Number'
ws['J1'] = 'Info'

ws.column_dimensions['A'].border = border_column
ws.column_dimensions['B'].border = border_column
ws.column_dimensions['C'].border = border_column
ws.column_dimensions['D'].border = border_column
ws.column_dimensions['E'].border = border_column
ws.column_dimensions['F'].border = border_column
ws.column_dimensions['G'].border = border_column
ws.column_dimensions['H'].border = border_column
ws.column_dimensions['I'].border = border_column
ws.column_dimensions['J'].border = border_column

ws.column_dimensions['A'].width = 45
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 15
ws.column_dimensions['I'].width = 15
ws.column_dimensions['J'].width = 40

ws['A1'].border = border_top_cell
ws['B1'].border = border_top_cell
ws['C1'].border = border_top_cell
ws['D1'].border = border_top_cell
ws['E1'].border = border_top_cell
ws['F1'].border = border_top_cell
ws['G1'].border = border_top_cell
ws['H1'].border = border_top_cell
ws['I1'].border = border_top_cell
ws['J1'].border = border_top_cell

print(incorrect_items, incorrect_names)

file_names_dict = {}

for name, item in zip(incorrect_names, incorrect_items):
    if name in file_names_dict:
        item = file_names_dict[name] + ', ' + item
    file_names_dict.update({name: item})

print(file_names_dict)
incorrect_names = list(file_names_dict.keys())
incorrect_items = list(file_names_dict.values())

master_dict = {}

i = 0
tup = []
for r, name in enumerate(incorrect_names, start=2):
    info = 'N/A'
    if name.count('-') > 6:
        project_code, originator, zone_system, level, type_doc, role, number, info = name.split('-')
        info = info[:info.find('.')]
    else:
        project_code, originator, zone_system, level, type_doc, role, number = name.split('-')
        number = number[:number.find('.')]

    master_dict[name] = {'project_code': project_code, 'originator': originator, 'zone_system': zone_system, 'level': level, 'type_doc': type_doc, 'role': role, 'number': number, 'info': info}
    d = (name, project_code, originator, zone_system, level, type_doc, role, number, info)

    ws[f'A{r}'] = name
    ws[f'B{r}'] = incorrect_items[i]
    ws[f'C{r}'] = project_code
    ws[f'D{r}'] = originator
    ws[f'E{r}'] = zone_system
    ws[f'F{r}'] = level
    ws[f'G{r}'] = type_doc
    ws[f'H{r}'] = role
    ws[f'I{r}'] = number
    ws[f'J{r}'] = info
    cell1 = ws[f'A{r}']
    cell2 = ws[f'B{r}']
    cell3 = ws[f'C{r}']
    cell4 = ws[f'D{r}']
    cell5 = ws[f'E{r}']
    cell6 = ws[f'F{r}']
    cell7 = ws[f'G{r}']
    cell8 = ws[f'H{r}']
    cell9 = ws[f'I{r}']
    cell10 = ws[f'J{r}']

    cell2.font = cell2.font.copy(color='FF0000')

    cell1.border = border_column
    cell2.border = border_column
    cell3.border = border_column
    cell4.border = border_column
    cell5.border = border_column
    cell6.border = border_column
    cell7.border = border_column
    cell8.border = border_column
    cell9.border = border_column
    cell10.border = border_column


    cells = [cell1, cell2, cell3, cell4, cell5, cell6, cell7, cell8, cell9, cell10]
    tup.append(incorrect_items[i])

    print(incorrect_items_code)
    print(incorrect_items)

    for cell in cells:
        if len(cell.value.split(', ')) <= 1:
            print(incorrect_items_code[i])
            print('HERE', cell.value + incorrect_items_code[i].split('_')[1], incorrect_items_code[i])
        if cell.value + incorrect_items_code[i].split('_')[1] == incorrect_items_code[i]:
            cell.font = cell.font.copy(color='FF0000')
        else:
            cell.font = cell.font.copy(color='FF0000')

    # for item in incorrect_items_code:
    #     incorrect_item, er_code = item.split('_')

    # for cell in cells:
    #     if len(cell.value.split(',')) > 1 and cell.value in incorrect_items:
    #         cell.font = cell.font.copy(color='FF0000')
    #     if check_incorrect_item(incorrect_items, incorrect_items[i], cell.value, list_files_names, tup):
    #         cell.font = cell.font.copy(color='FF0000')
    i += 1
wb.save('Incorrect_Naming2.xlsx')

print(master_dict)
