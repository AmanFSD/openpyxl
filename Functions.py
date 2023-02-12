import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side

border_column = Border(right=Side(style='thin'))
border_top_cell = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

def get_files():
    folder_path = input('Folder Path? >')
    os.chdir(rf'{folder_path}')
    files = os.listdir()
    return files


def map_files_with_keys(files):
    files_dict = {}
    for index, file in enumerate(files, start=1):
        files_dict.update({index: file})
        print(f'{index}: {file}')
    return files_dict


def extracting_files_names_from_excel(files_dict, file_number):
    df = pd.read_excel(files_dict[file_number], sheet_name='Report Content')

    list_files_name = list(df.Name)
    return list_files_name


def find_incorrect_file_name(file_name, incorrect_names, incorrect_items):
    extension = file_name[file_name.find('.'):]

    if '-' not in file_name or '_' in file_name[:21] or file_name[4] != '-' or file_name[8] != '-' or file_name[11] != '-' or file_name[14] != '-' or file_name[17] != '-' or file_name[21] != '-':
        print('incorrect files names, missing \"-\".')
        incorrect_names.append(file_name)
    else:
        names = file_name.split('-')
        names[-1] = names[-1][:-4]
        if names[0] not in ['VCE1']:
            incorrect_items.append(names[0])
            incorrect_names.append(file_name)
        if extension == '.rvt' and names[4] not in ['M3']:
            incorrect_items.append(names[4])
            incorrect_names.append(file_name)
        elif extension == '.nwd' and names[4] not in ['M3', 'CR']:
            incorrect_items.append(names[4])
            incorrect_names.append(file_name)
        elif extension == '.nwc' and names[4] not in ['M3', 'CR']:
            incorrect_items.append(names[4])
            incorrect_names.append(file_name)
        elif extension == '.pdf' and names[4] not in ['M2', 'DR']:
            incorrect_items.append(names[4])
            incorrect_names.append(file_name)
        elif extension == '.dwg' and names[4] not in ['M2', 'DR']:
            incorrect_items.append(names[4])
            incorrect_names.append(file_name)

def create_modify_excel_sheet():
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'File Names'
    ws['B1'] = 'Incorrect Item'

    ws.column_dimensions['A'].border = border_column
    ws.column_dimensions['B'].border = border_column

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15

    ws['A1'].border = border_top_cell
    ws['B1'].border = border_top_cell
    return wb, ws


def import_data_in_excel(ws, incorrect_names, incorrect_items):
    i = 0
    for r, name in enumerate(incorrect_names, start=2):
        ws[f'A{r}'] = name
        ws[f'B{r}'] = incorrect_items[i]
        cell1 = ws[f'A{r}']
        cell2 = ws[f'B{r}']
        cell2.font = cell2.font.copy(color='FF0000')
        cell1.border = border_column
        cell2.border = border_column
        i += 1